#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
為結構化題庫產生「難度」與「主題標籤」初稿（自動、未經人工審核）。
方法：以題目特徵做啟發式評分 → 依全題庫百分位分成 易/中/難；
      標籤＝章節主題 + 關鍵字命中。結果僅供人工審核前的初稿。
用法： python tag_difficulty.py 結構化題庫.xlsx
會就地更新該檔的「難度」「標籤」兩欄，其餘分頁保留。
"""
import sys, re
import openpyxl

PATH = sys.argv[1] if len(sys.argv) > 1 else '公共工程品管(機電)題庫_結構化_115年1月起適用.xlsx'

# 章節 → 主題標籤（完整章節名為鍵，避開「單元二」兩組相撞）
CHAP_TAG = {
 '政府採購全生命週期概論':'政府採購',
 '第一章 公共工程施工品質管理制度理念與導入':'品管制度',
 '第二章 公共工程管理相關法規概要':'工程法規',
 '第三章 公共工程履約管理':'履約管理',
 '第四章 永續公共工程':'永續工程',
 '第五章 工程倫理':'工程倫理',
 '第一章 監造計畫與品質計畫指導':'監造與品質計畫',
 '第二章 統計分析方法與應用':'統計分析',
 '第三章 品質分析方法與應用':'品質分析',
 '第四章 工程進料檢驗與管制':'進料檢驗',
 '第五章 施工管制與檢驗':'施工管制',
 '第六章 工程品質稽核':'品質稽核',
 '第一章 電氣設備工程品質管理實務':'電氣設備',
 '第二章 弱電(含中央監控)設備工程品質管理實務':'弱電與中央監控',
 '第三章 給排水衛生設備工程品質管理實務':'給排水衛生',
 '第四章 消防設備及空調設備工程品質管理實務':'消防與空調',
 '第五章 電梯設備工程品質管理實務':'電梯設備',
 '第六章 建築工程之界面整合':'界面整合',
}
# 關鍵字 → 次要標籤
KW_TAG = [
 ('契約','契約管理'),('物價','物價調整'),('綠建築','綠建築'),('倫理','工程倫理'),
 ('監造','監造'),('驗收','驗收'),('estimate',''),('進料','進料檢驗'),
 ('管制圖','品管手法'),('直方圖','品管手法'),('柏拉圖','品管手法'),('特性要因','品管手法'),
 ('KJ','品管手法'),('親和圖','品管手法'),('抽樣','抽樣檢驗'),('標準差','統計量'),
 ('常態','統計量'),('變異','統計量'),('職業安全','職業安全衛生'),('安全衛生','職業安全衛生'),
 ('消防','消防'),('空調','空調'),('電梯','電梯'),('接地','電氣安全'),('絕緣','電氣安全'),
 ('發電機','發電設備'),('不斷電','不斷電系統'),('UPS','不斷電系統'),('給水','給排水'),
 ('排水','給排水'),('採購','政府採購'),('查核','工程查核'),('稽核','品質稽核'),
 ('缺工','營建人力'),('移工','營建人力'),('環保','環境保護'),('廢棄物','環境保護'),
]

NEG = re.compile(r'(何者非|下列何者不|何者錯|不正確|錯誤|不屬|非屬|不包含|不需)')
ALLOF = re.compile(r'以上皆|皆是|皆非|皆可')
MULTI = re.compile(r'1[\.\、].*2[\.\、].*3[\.\、]')
LAW = re.compile(r'(法|條|要點|規範|標準|辦法|準則|規則|規定|手冊|基準)')

def score(stem, opts):
    s = 0.0
    s += min(len(stem), 120) / 40.0
    s += min(sum(len(o) for o in opts), 120) / 60.0
    if NEG.search(stem): s += 1.2
    if ALLOF.search(' '.join(opts)): s += 0.5
    if MULTI.search(stem): s += 1.8
    s += min(len(re.findall(r'\d', stem)), 12) * 0.12
    s += min(len(LAW.findall(stem)), 8) * 0.18
    return s

wb = openpyxl.load_workbook(PATH)
ws = wb['題庫'] if '題庫' in wb.sheetnames else wb.worksheets[0]
hdr = {c.value: i for i, c in enumerate(ws[1]) if c.value}
ci_chap, ci_stem = hdr['章節'], hdr['題目']
ci_opts = [hdr['選項'+L] for L in ['A','B','C','D']]
ci_diff, ci_tag = hdr['難度'], hdr['標籤']

rows = list(ws.iter_rows(min_row=2))
scored = []
for row in rows:
    stem = str(row[ci_stem].value or '')
    opts = [str(row[c].value or '') for c in ci_opts]
    scored.append(score(stem, opts))

order = sorted(range(len(scored)), key=lambda i: scored[i])
rank = [0]*len(scored)
for pos, i in enumerate(order):
    rank[i] = pos / max(1, len(scored)-1)   # 0..1 百分位

def bucket(p):
    if p < 0.33: return '易'
    if p < 0.78: return '中'
    return '難'

dist = {'易':0,'中':0,'難':0}
for i, row in enumerate(rows):
    stem = str(row[ci_stem].value or '')
    chap = str(row[ci_chap].value or '')
    d = bucket(rank[i])
    dist[d] += 1
    tags = []
    primary = CHAP_TAG.get(chap)
    if primary: tags.append(primary)
    for kw, tag in KW_TAG:
        if tag and kw in stem and tag not in tags:
            tags.append(tag)
        if len(tags) >= 3: break
    row[ci_diff].value = d
    row[ci_tag].value = '、'.join(tags)

wb.save(PATH)
print('難度分布：', dist)
print('已更新：', PATH)
