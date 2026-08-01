#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
匯入結構化題庫 Excel 到 PostgreSQL（schema_mvp.sql）

用法：
  1. 先建好資料庫並套用 schema：  psql -d quizdb -f schema_mvp.sql
  2. 設定連線（擇一）：
       export DATABASE_URL=postgresql://user:pass@host:5432/quizdb
     或使用標準 PG 環境變數 PGHOST/PGPORT/PGUSER/PGPASSWORD/PGDATABASE
  3. 執行：
       python import_questions.py "公共工程品管(機電)題庫_結構化_115年1月起適用.xlsx"

特性：
  - 自動建立 units / chapters（依「單元」「章節」，保留兩組「單元二」的完整章節名）
  - question_key = <BANK_CODE>-<原始題號4碼>，跨題庫版本穩定（錯題本／學習追蹤用）
  - 選項寫入 question_options，正解標記 is_correct
  - 可重複執行：重跑會先清除該版本既有內容再重灌（idempotent）
"""
import os, sys, argparse
import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# 對應 schema_mvp.sql 的種子 UUID（平台租戶 / 題庫 / 版本）
PLATFORM_TENANT = '00000000-0000-0000-0000-000000000001'
BANK_ID         = '00000000-0000-0000-0000-0000000000b1'
BANK_VERSION_ID = '00000000-0000-0000-0000-0000000000c1'
BANK_CODE       = 'me'   # 機電；question_key 前綴

DIFF_MAP   = {'易':'easy','中':'medium','難':'hard','簡單':'easy','困難':'hard'}
STATUS_MAP = {'草稿':'draft','待審核':'pending','已發布':'published',
              '有疑義':'disputed','已修訂':'revised','已停用':'retired'}

def col_index(ws):
    """讀表頭，回傳 {欄名: 欄索引(0-based)}"""
    header = [c.value for c in ws[1]]
    return {name: i for i, name in enumerate(header) if name is not None}

def load_rows(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['題庫'] if '題庫' in wb.sheetnames else wb.worksheets[0]
    idx = col_index(ws)
    need = ['題號','單元','章節','題型','題目','選項A','選項B','選項C','選項D','正確答案']
    missing = [c for c in need if c not in idx]
    if missing:
        sys.exit(f'Excel 缺少欄位：{missing}')
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[idx['題目']] is None:
            continue
        def g(name):
            i = idx.get(name)
            return r[i] if i is not None and i < len(r) else None
        rows.append({
            'no': int(g('題號')),
            'unit': (g('單元') or '概論'),
            'chapter': g('章節'),
            'type': g('題型') or '單選題',
            'stem': g('題目'),
            'opts': {L: g('選項'+L) for L in ['A','B','C','D','E'] if g('選項'+L)},
            'ans': (str(g('正確答案')).strip().upper() if g('正確答案') else None),
            'difficulty': DIFF_MAP.get(str(g('難度')).strip()) if g('難度') else None,
            'source': g('題目來源'),
            'status': STATUS_MAP.get(str(g('狀態')).strip(), 'published') if g('狀態') else 'published',
            'tags': [t.strip() for t in str(g('標籤')).replace('，','、').split('、') if t.strip()] if g('標籤') else [],
            'explanation': g('解析'),
            'legal_basis': g('法規依據'),
        })
    return rows

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('xlsx', nargs='?',
                    default='公共工程品管(機電)題庫_結構化_115年1月起適用.xlsx')
    args = ap.parse_args()

    rows = load_rows(args.xlsx)
    print(f'讀到 {len(rows)} 題')

    dsn = os.environ.get('DATABASE_URL')
    conn = psycopg2.connect(dsn) if dsn else psycopg2.connect()
    conn.autocommit = False
    cur = conn.cursor()

    # 確保 種子（若已由 schema 建立則略過）
    cur.execute("INSERT INTO tenants(id,type,name) VALUES(%s,'platform','平台（自有品牌）') ON CONFLICT DO NOTHING", (PLATFORM_TENANT,))
    cur.execute("""INSERT INTO question_banks(id,owner_tenant_id,name,exam_category,is_shared,status)
                   VALUES(%s,%s,'公共工程品質管理（機電）','公共工程品管',true,'active') ON CONFLICT DO NOTHING""",
                (BANK_ID, PLATFORM_TENANT))
    cur.execute("""INSERT INTO bank_versions(id,bank_id,version_label,effective_date,status)
                   VALUES(%s,%s,'115年1月起適用','2026-01-01','published') ON CONFLICT DO NOTHING""",
                (BANK_VERSION_ID, BANK_ID))

    # 清除此版本既有內容（可重跑）
    cur.execute("DELETE FROM questions WHERE bank_version_id=%s", (BANK_VERSION_ID,))
    cur.execute("DELETE FROM chapters WHERE unit_id IN (SELECT id FROM units WHERE bank_version_id=%s)", (BANK_VERSION_ID,))
    cur.execute("DELETE FROM units WHERE bank_version_id=%s", (BANK_VERSION_ID,))

    # 建 units（依首次出現順序）
    unit_id = {}
    for r in rows:
        if r['unit'] not in unit_id:
            cur.execute("INSERT INTO units(bank_version_id,name,sort_order) VALUES(%s,%s,%s) RETURNING id",
                        (BANK_VERSION_ID, r['unit'], len(unit_id)))
            unit_id[r['unit']] = cur.fetchone()[0]

    # 建 chapters（unit 內依首次出現順序；用完整章節名，避免「單元二」兩組相撞）
    chap_id = {}
    for r in rows:
        key = (r['unit'], r['chapter'])
        if key not in chap_id:
            cur.execute("INSERT INTO chapters(unit_id,name,sort_order) VALUES(%s,%s,%s) RETURNING id",
                        (unit_id[r['unit']], r['chapter'], len([k for k in chap_id if k[0]==r['unit']])))
            chap_id[key] = cur.fetchone()[0]

    # tags（平台共用；tenant_id=NULL）
    tag_id = {}
    def get_tag(name):
        if name in tag_id: return tag_id[name]
        cur.execute("""INSERT INTO tags(tenant_id,name) VALUES(NULL,%s)
                       ON CONFLICT (tenant_id,name) DO UPDATE SET name=EXCLUDED.name RETURNING id""", (name,))
        tag_id[name] = cur.fetchone()[0]
        return tag_id[name]

    qtype = {'單選題':'single','複選題':'multiple','是非題':'true_false'}
    n_opt = 0
    for r in rows:
        qkey = f"{BANK_CODE}-{r['no']:04d}"
        cur.execute("""INSERT INTO questions(bank_version_id,chapter_id,question_key,original_no,type,stem,difficulty,source,status)
                       VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (BANK_VERSION_ID, chap_id[(r['unit'],r['chapter'])], qkey, r['no'],
                     qtype.get(r['type'],'single'), r['stem'], r['difficulty'], r['source'], r['status']))
        qid = cur.fetchone()[0]
        opt_vals = [(qid, L, r['opts'][L], (L==r['ans']), i)
                    for i, L in enumerate(sorted(r['opts']))]
        execute_values(cur,
            "INSERT INTO question_options(question_id,label,content,is_correct,sort_order) VALUES %s", opt_vals)
        n_opt += len(opt_vals)
        if r['explanation'] or r['legal_basis']:
            cur.execute("""INSERT INTO question_explanations(question_id,explanation,legal_basis,ai_generated,reviewed)
                           VALUES(%s,%s,%s,false,false)
                           ON CONFLICT (question_id) DO UPDATE SET explanation=EXCLUDED.explanation""",
                        (qid, r['explanation'], r['legal_basis']))
        for t in r['tags']:
            cur.execute("INSERT INTO question_tags(question_id,tag_id) VALUES(%s,%s) ON CONFLICT DO NOTHING",
                        (qid, get_tag(t)))

    conn.commit()
    print(f'完成：units={len(unit_id)} chapters={len(chap_id)} questions={len(rows)} options={n_opt} tags={len(tag_id)}')
    cur.close(); conn.close()

if __name__ == '__main__':
    main()
